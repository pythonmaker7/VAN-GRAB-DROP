#!/usr/bin/env python3

import copy
import openpyxl
import json

tables = openpyxl.load_workbook('../data/tables.xlsx')
hubs = tables[tables.sheetnames[0]]
garms = tables[tables.sheetnames[1]]
orders_file = open('../data/orders.json')
route_file = open('../data/route.json')


# 1         all_hubs

all_hubs = {}

for i in range(2, hubs.max_row + 1):
    
    hub_id = hubs.cell(row=i, column=1).value
    hub_name = hubs.cell(row=i, column=2).value
    hub_latitude = hubs.cell(row=i, column=3).value
    hub_longtitude = hubs.cell(row=i, column=4).value
    hub_cap = hubs.cell(row=i, column=5).value
        
    all_hubs[hub_id] = [hub_name, hub_cap, []]
    

for hub_key, garms_in_hub in all_hubs.items():
    
    for i in range(2, garms.max_row+1):
        
        hub_id = garms.cell(row=i, column=3).value
        garm_id = garms.cell(row=i, column=1).value
        
        if hub_id == hub_key:
            garms_in_hub[2].append(garm_id)
    
#for i,e in all_hubs.items():
#   print(i,e)
    
# 2         item_hub_order
            
item_hub_order = json.load(orders_file)

# 2.2
item_hub_order = {
    int(key):int(value) for key,value in item_hub_order.items()
}

# 3        item_hub_move

item_hub_move = copy.deepcopy(item_hub_order)

for item_id in list(item_hub_order.keys()):
    for hub_id in all_hubs:
        if hub_id == item_hub_order[item_id]:
            if item_id in all_hubs[hub_id][2]:
                item_hub_move.pop(item_id) 

# 4         item_hub_stock

item_hub_stock = {}

for i in range(2,garms.max_row+1):    #4.2

    garm_id = garms.cell(row=i, column=1).value
    garm_ids = item_hub_move.keys()
    hub_id = garms.cell(row=i, column=3).value 
    
    if garm_id in garm_ids:   #4.3
        item_hub_stock[garm_id] = hub_id

# 5         item_hub_hub

item_hub_hub = {}

for item in list(item_hub_stock.keys()):
    item_hub_hub[item] = [item_hub_stock[item], item_hub_order[item]]

# 6         hub_route

hub_route = json.load(route_file)

# 7         hub_take_leave

hub_take_leave = {}
take = []
leave = []

for hub_id, hub_data in all_hubs.items():
    
    for item,v in item_hub_move.items():
        if v == hub_id:
            leave.append(item)
        if item in hub_data[2]:
            take.append(item)
            
    hub_take_leave[hub_id] = [take, leave]
    take = []
    leave = []

# 8         station_take_leave
    
station_take_leave = {}
van = []
take = []
leave = []
mid_station = int((len(hub_route)-1)/2) + 1

# prwta ftiaxnw tis prwtes staseis (prwtes fores) pou einai ta misa+1
for station in hub_route[:mid_station]:
    
    # ta pairnw ola
    take = hub_take_leave[station][0]
    
    for i in take:
        van.append(i)
        
    # ελεγχω αν υπαρχουν στο βαν.
    for i in hub_take_leave[station][1]:
        
        #αν ναι τοτε τα αφηνω
        if i in van:
            leave.append(i)
            van.remove(i)
            
    station_take_leave[station,'#1'] = [take, leave]
    take = []
    leave = []
    
# kai twra ftiaxnw tis alles mises (defteres staseis)
for station in hub_route[mid_station:]:
    
    # δεν παιρνω τιποτα γιατι εχω ηδη περασει
    
    # ελεγχω αν υπαρχουν στο βαν.
    for i in hub_take_leave[station][1]:
        
        #αν ναι τοτε τα αφηνω
        if i in van:
            leave.append(i)
            van.remove(i)
            
    station_take_leave[station,'#2'] = [take, leave]
    take = []
    leave = []

# OUTPUT
# OUTPUT
# OUTPUT
# OUTPUT
# OUTPUT
# OUTPUT
# OUTPUT
# OUTPUT
    

for station, take_leave in station_take_leave.items():
        
    hub_id = station[0]
    hub_name = all_hubs[station[0]][0]
    visit = station[1]
    
    # hide nothing stations
    if take_leave[0] == [] and take_leave[1] == []:
        continue
    
    print(hub_id, hub_name, visit)
    print('GRAB ', end="| ")
    print(*take_leave[0], sep=', ')
    print('DROP ', end="| ")
    print(*take_leave[1], sep=', ')
    print('')